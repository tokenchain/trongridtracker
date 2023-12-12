#!/usr/bin/env python3
# coding: utf-8
import glob
import json
import os

from .builder import build_bot_tpl
from .mistapi import MistAcquireDat, get_mist_graph_api
import pandas as pd
from pandas import DataFrame
import openpyxl as ox
from core.common.utils import FolderBase, ReadCombinedFile, find_key


class MistAnalysis(FolderBase):
    """
    https://graphviz.org/doc/info/shapes.html

    """

    def __init__(self, project: str):
        self.by_project_name(project)
        self.handle_address = ""
        self.reset_metadata()
        self.rendered = False
        self.scope = 500
        self.dot = build_bot_tpl(
            "collection",
            self.scope
        )
        self._main_address = ""
        self.edges = 0
        self.use_from = False
        self.use_to = False
        self.is_independence = False
        self.enable_sidenote = False
        self.api = MistAcquireDat(project)
        self.api.folder = self.mist_folder
        self.api.inputfolder = self.inputfolder
        self.project_name = project

    def doIndependChart(self):
        self.is_independence = True
        return self

    def reset_metadata(self):
        self.metadata = {
            "LINK": {},
            "IDS": {},
            "USE_NODE": [],
            "FROM_LIST": [],
        }

    def setThreadHoldUSD(self, n):
        self.scope = n
        return self

    def setUseFrom(self):
        self.use_from = True
        self.use_to = False
        return self

    def setUseTo(self):
        self.use_from = False
        self.use_to = True
        return self

    def setEnableSideNote(self):
        self.enable_sidenote = True
        return self

    def startPlot(self):
        file_pattern = "*.json"  # Replace with your specific file name pattern
        # Use glob to search for files with the specified name pattern
        # search = os.path.join(os.path.dirname(__file__), self.mist_folder, file_pattern)
        search = os.path.join(self.mist_folder, file_pattern)
        file_list = glob.glob(search)
        if self.is_independence is False:
            # Loop over each file and perform some operation
            for file_path in file_list:
                # file_name = os.path.basename(file_path)
                self._inside(file_path, self.scope)
            self.end()
        else:
            y = 0
            for file_x in file_list:
                self.reset_metadata()
                self.dot = build_bot_tpl(
                    "collection",
                    self.scope
                )
                self._inside(file_x, self.scope)
                self.end_with_k(y)
                y += 1

    def start_develop_source_sheet(self, file_name: str):
        path = os.path.join(self.chart_folder, f"{file_name}.xlsx")
        file_pattern = "*.json"  # Replace with your specific file name pattern
        # Use glob to search for files with the specified name pattern
        # search = os.path.join(os.path.dirname(__file__), self.mist_folder, file_pattern)
        search = os.path.join(self.mist_folder, file_pattern)
        file_list = glob.glob(search)
        # Loop over each file and perform some operation
        for file_path in file_list:
            file_name = os.path.basename(file_path)
            self._personal_note(file_path, self.scope)
        # self.end()
        self._excel_factory(path)

    def _personal_note(self, file, scope):
        self._main_address = ""
        with open(file, newline='') as f:
            self.rf = json.loads(f.read())
            # render_node = True
            if "graph_dic" not in self.rf:
                return
            text = f"下图显示了所有被 {scope} USDT 或以下忽略的记录交易"
            with self.dot.subgraph(name='legend') as c:
                c.attr(color='blue')
                c.node_attr['style'] = 'filled'
                c.attr(label=text)

            nodes = self.rf["graph_dic"]["node_list"]
            edges = self.rf["graph_dic"]["edge_list"]

            for v in nodes:
                _font_colr = "slateblue4"
                _id = v["id"]
                _add = v["addr"]
                _lab = ""

                if "shape" in v:
                    _shape = v["shape"]
                    if "star" in _shape:
                        self._main_address = _id

                else:
                    _shape = "box"

                if _shape == "circularImage":
                    _shape = "box"

                if _add not in self.metadata:
                    self.metadata[_add] = _id
                else:
                    old_id = self.metadata[_add]
                    self.metadata["LINK"][_id] = old_id
                    _id = old_id

                    # render_node = False

                if "..." in v["label"]:
                    _lab = v["addr"]
                    _color = "springgreen2"
                    side_note = self.api.overview(_add)
                    _lab = _lab + side_note
                    _font_colr = "white"
                    _shape = "folder"

                if "color" in v:
                    _color = v["color"]

                self.metadata["IDS"][_id] = {
                    "shape": _shape,
                    "fillcolor": _color,
                    "style": "filled",
                    "fontcolor": _font_colr,
                    "label": _lab,
                    "address": v["addr"]
                }

            for v in edges:
                if v["val"] < scope:
                    continue
                if "color" in v:
                    _color = v["color"]["color"]
                else:
                    _color = "red"
                count = len(v["tx_hash_list"])
                __label = v["label"]
                __label_arrow = f"{count}筆,{__label}"
                self.dot.edge(
                    self.getId(v["from"]),
                    self.getId(v["to"]),
                    color=_color,
                    label=__label_arrow
                )

                if self._main_address == self.getId(v["to"]) and self.use_from:
                    from_id = self.getId(v["from"])
                    address_from = self.metadata["IDS"][from_id]["address"]
                    # print("record -> excel")
                    self.metadata["FROM_LIST"].append({
                        "info": self.api.overviewdict(address_from),
                        "address": address_from,
                        "spent_count": count,
                        "contribute": __label
                    })

                if self._main_address == self.getId(v["from"]) and self.use_to:
                    from_id = self.getId(v["to"])
                    address_from = self.metadata["IDS"][from_id]["address"]
                    # print("record -> excel")

                    self.metadata["FROM_LIST"].append({
                        "info": self.api.overviewdict(address_from),
                        "address": address_from,
                        "spent_count": count,
                        "contribute": __label
                    })

                self.edges += 1

            for _id in self.metadata["USE_NODE"]:
                # if render_node:
                ob = self.metadata["IDS"][_id]
                self.dot.node(
                    _id,
                    shape=ob["shape"],
                    fillcolor=ob["fillcolor"],
                    style="filled",
                    fontcolor=ob["fontcolor"],
                    label=ob["label"],
                    constraint="false"
                )

    def _inside(self, file, scope):
        with open(file, newline='') as f:
            self.rf = json.loads(f.read())
            print(file)
            # render_node = True
            if "graph_dic" not in self.rf:
                return
            text = f"下图显示了所有被 {scope} USDT 或以下忽略的记录交易"

            with self.dot.subgraph(name='legend') as c:
                c.attr(color='blue')
                c.node_attr['style'] = 'filled'
                c.attr(label=text)

            nodes = self.rf["graph_dic"]["node_list"]
            edges = self.rf["graph_dic"]["edge_list"]

            for v in nodes:
                _font_colr = "black"
                _id = v["id"]
                _add = v["addr"]

                if "shape" in v:
                    _shape = v["shape"]
                else:
                    _shape = "box"

                if _shape == "circularImage":
                    _shape = "box"

                if _add not in self.metadata:
                    self.metadata[_add] = _id
                else:
                    old_id = self.metadata[_add]
                    self.metadata["LINK"][_id] = old_id
                    _id = old_id
                    # render_node = False

                if "..." in v["label"]:
                    _lab = v["addr"]
                    _color = "lightgreen"
                    if "star" in _shape:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note
                        _font_colr = "white"
                        _shape = "folder"
                        _color = "darkslategrey"
                else:
                    _color = "lightcoral"
                    _lab = f"[{v['label']}]\n{v['addr']}"
                    if "star" in _shape:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note
                        _font_colr = "white"
                        _shape = "folder"
                        _color = "darkslategrey"

                    if self.enable_sidenote:
                        side_note = self.api.overview(_add)
                        _lab = _lab + side_note

                # if "color" in v:
                #    _color = v["color"]

                self.metadata["IDS"][_id] = {
                    "shape": _shape,
                    "fillcolor": _color,
                    "style": "filled",
                    "fontcolor": _font_colr,
                    "label": _lab,
                }

            for v in edges:
                if v["val"] < scope:
                    continue
                if "color" in v:
                    _color = v["color"]["color"]
                else:
                    _color = "red"
                count = len(v["tx_hash_list"])
                __label = v["label"]
                self.dot.edge(
                    self.getId(v["from"]),
                    self.getId(v["to"]),
                    color=_color,
                    label=f"{count}筆,{__label}",
                    labeldistance='1.2',
                    labelangle='60',
                )

                self.edges += 1

            for _id in self.metadata["USE_NODE"]:
                # if render_node:
                ob = self.metadata["IDS"][_id]
                self.dot.node(
                    _id,
                    shape=ob["shape"],
                    fillcolor=ob["fillcolor"],
                    style="filled",
                    fontcolor=ob["fontcolor"],
                    label=ob["label"]
                )

    def getId(self, _idx):
        if _idx in self.metadata["LINK"]:
            use_alternative = self.metadata["LINK"][_idx]
        else:
            use_alternative = _idx

        if use_alternative not in self.metadata["USE_NODE"]:
            self.metadata["USE_NODE"].append(use_alternative)

        return use_alternative

    def end(self):
        self.dot.render(
            filename=f"{self.project_name}.{self.scope}",
            directory=self.chart_folder).replace('\\', '/')

    def end_with_k(self, k: int):
        self.dot.render(
            filename=f"{self.project_name}.{self.scope}.{k}",
            directory=self.chart_folder).replace('\\', '/')

    def setName(self, name: str):
        self.project_name = name

    def _readDF(self, excel_file: str) -> DataFrame:
        data = pd.read_excel(excel_file)
        df = pd.DataFrame(data, columns=self._excel_header)
        return df

    def _excel_factory(self, excel_file: str) -> "MistAnalysis":

        self._excel_header = [
            # 2         3         4       5        6         7     8          9            10
            "address", "start", "end", "income", "outflow", "net", "tx", "spent count", "contribution"
        ]
        sheet_name = ""

        if self.use_to is True:
            sheet_name = "income"

        if self.use_from is True:
            sheet_name = "expense"

        try:
            _df = self._readDF(excel_file)
            wb = ox.load_workbook(excel_file)
            ws = wb[sheet_name]
            startrow: int = 0
            original_rows = len(self.metadata["FROM_LIST"])
            for row in range(0, original_rows):
                if row >= startrow:
                    rT = row + 2
                    package_info = self.metadata["FROM_LIST"][row]["info"]
                    address = self.metadata["FROM_LIST"][row]["address"]
                    spent_count = self.metadata["FROM_LIST"][row]["spent_count"]
                    contribute = self.metadata["FROM_LIST"][row]["contribute"]

                    ws.cell(row=rT, column=1).value = address
                    ws.cell(row=rT, column=2).value = package_info["first_tx_time"]
                    ws.cell(row=rT, column=3).value = package_info["last_tx_time"]
                    ws.cell(row=rT, column=4).value = package_info["total_received"]
                    ws.cell(row=rT, column=5).value = package_info["total_spent"]
                    ws.cell(row=rT, column=6).value = package_info["balance"]
                    ws.cell(row=rT, column=7).value = package_info["tx_count"]
                    ws.cell(row=rT, column=8).value = spent_count
                    ws.cell(row=rT, column=9).value = contribute
                    print(address)
                    # for ch in range(1, original_cols + 1):
                    # ws.cell(row=rT, column=ch).fill = pf
                    #    pass

            """
            for row in range(0, original_rows):  # For each row in the dataframe
                for col in range(0, original_cols):  # For each column in the dataframe
                    if row >= startrow:
                        val_old = _df.iat[row, col]
                        key = self._excel_header[col]
                        if col == 0:
                            rT = row + 2
                            ws.cell(row=rT, column=2).value = pass_discord
                            ws.cell(row=rT, column=4).value = pass_email
                            ws.cell(row=rT, column=5).value = token_p
                            ws.cell(row=rT, column=9).value = banned
                            ws.cell(row=rT, column=10).value = last_operation
                            ws.cell(row=rT, column=11).value = bind_phone
                            # print(f"Pop =|> {rT} {val_old} {k1v1} {k1v2} {k1v3}")
            """

            wb.save(excel_file)

        except FileNotFoundError:
            pass

        return self

    def acquireFromAddress(self, addresses: list):
        for x in addresses:
            self.api.save(x)

    def acquireByFile(self, file_name: str):
        addresses = []
        path = os.path.join(self.inputfolder, file_name)
        if os.path.isfile(path) is False:
            print("the address file is not found under the inputs folder.")
            exit(2)

        xexe = open(path, "r")
        print("open file x", file_name)
        lines = xexe.readlines()
        xexe.close()

        lines = [h.replace("\n", "") for h in lines]
        addresses += lines

        for x in addresses:
            print(x)
            if x[:2] == "0x":
                self.api.save(x)