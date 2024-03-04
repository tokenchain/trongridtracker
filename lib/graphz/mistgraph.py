#!/usr/bin/env python3
# coding: utf-8
from .builder import build_bot_tpl
from .mistapi import *
import openpyxl as ox
from lib.common.utils import *

import statistics


class LayerManager():
    """
    ma.acquireFromAddress([
        "TEZVBjomubVSn41hvenJ8KyHEzEZCeuC7B"
    ])
    ma.setName("christmas-pdd")
    ma.lineThickness()
    ma.setThreadHoldUSD(bar_small).setEnableSideNote().startPlot()

    """

    def __init__(self):
        self.current_depth = 0
        self.analysis_depth = 0
        self.layers_address = []
        self.layers_address.append([])

    def set_depth(self, x: int):
        self.analysis_depth = x

    def add_first_layer(self, addresses: list):
        self.layers_address[0] = addresses

    def next_depth(self) -> bool:
        if self.current_depth < self.analysis_depth:
            total_addresses = len(self.layers_address[self.current_depth])
            self.current_depth += 1
            # self.layers_address[self.current_depth] = []
            self.layers_address.append([])
            print("NOW ANNOUNCE THE NEXT LAYER!!!!! addr count", total_addresses)
            return True
        else:
            return False

    def get_current_layer_address(self) -> list:
        return self.layers_address[self.current_depth]

    def dump_layer_addresses(self, add1: str, add2: str, common1: bool, common2: bool):
        path_layer = self.layers_address[self.current_depth]
        if add1 not in path_layer and common1 is True:
            path_layer.append(add1)
        if add2 not in path_layer and common2 is True:
            path_layer.append(add2)
        self.layers_address[self.current_depth] = path_layer


class MistAnalysis(FolderBase):
    """
    https://graphviz.org/doc/info/shapes.html

    """

    def __init__(self, project: str):
        if "/" in project or "\"" in project:
            print("malformed project name")
            exit(0)
        self.metadata = None
        self.by_project_name(project)
        self.handle_address = ""
        self.reset_metadata()
        self.rendered = False
        self.scope = 500
        self.dot = build_bot_tpl(
            "collection",
            self.scope
        )
        self._main_address = ""
        self.edges = 0
        self.use_from = False
        self.use_to = False
        self.is_independence = False
        self.enable_sidenote = False
        self._wire_thickness_requirement = False
        self.api = MistAcquireDat(project)
        self.api.folder = self.mist_folder
        self.api.inputfolder = self.inputfolder
        self.project_name = project
        self.layer = LayerManager()

    def doIndependChart(self):
        self.is_independence = True
        return self

    def reset_metadata(self):
        self.metadata = {
            "LINK": {},
            "IDS": {},
            "USE_NODE": [],
            "FROM_LIST": [],
            "MEAN": 0,
            "STD": 0
        }

    def setThreadHoldUSD(self, n):
        self.scope = n
        return self

    def setUseFrom(self):
        self.use_from = True
        self.use_to = False
        return self

    def setUseTo(self):
        self.use_from = False
        self.use_to = True
        return self

    def setEnableSideNote(self):
        self.enable_sidenote = True
        return self

    def startPlotDepth(self, x: int):
        self.layer.set_depth(x)
        layer_addresses = self.layer.get_current_layer_address()
        self.layer.next_depth()

        while True:
            for address in layer_addresses:
                if self.api.sql.has_mapping(address) is True:
                    data_raw = self.api.sql.read_mapping(address)
                    if data_raw is False:
                        continue
                    self._inside(data_raw, self.scope)
                else:
                    self.api.save(address)
                    data_raw = self.api.sql.read_mapping(address)
                    if data_raw is False:
                        continue
                    self._inside(data_raw, self.scope)

            layer_addresses = self.layer.get_current_layer_address()
            self._processing_layer(layer_addresses)
            if self.layer.next_depth() is False:
                break

        self._render_all_nodes()
        self.end()

    def startPlot(self):
        """
        this is the classical implementation
        """
        file_pattern = "*.json"  # Replace with your specific file name pattern
        # Use glob to search for files with the specified name pattern
        # search = os.path.join(os.path.dirname(__file__), self.mist_folder, file_pattern)
        search = os.path.join(self.mist_folder, file_pattern)
        data_raw = None
        file_list = glob.glob(search)
        if self.is_independence is False:
            # Loop over each file and perform some operation
            while True:
                for file_path in file_list:
                    # file_name = os.path.basename(file_path)
                    with open(file_path, newline='') as f:
                        data_raw = json.loads(f.read())
                        f.close()

                    self._inside(data_raw, self.scope)

                if self.layer.next_depth() is False:
                    break

            self._render_all_nodes()
            self.end()
        else:
            y = 0
            for file_xi in file_list:
                with open(file_xi, newline='') as f:
                    data_raw = json.loads(f.read())
                    f.close()

                self.reset_metadata()
                self.dot = build_bot_tpl(
                    "collection",
                    self.scope
                )
                self._inside(data_raw, self.scope, None)
                self._render_all_nodes()
                self.end_with_k(y)
                y += 1

    def start_develop_source_sheet(self, file_name: str, coin_type: str):
        path = os.path.join(self.chart_folder, f"{file_name}.xlsx")
        file_pattern = "*.json"  # Replace with your specific file name pattern
        # Use glob to search for files with the specified name pattern
        # search = os.path.join(os.path.dirname(__file__), self.mist_folder, file_pattern)
        search = os.path.join(self.mist_folder, file_pattern)
        file_list = glob.glob(search)
        # Loop over each file and perform some operation
        for file_path in file_list:
            # file_name = os.path.basename(file_path)
            self._personal_note(file_path, self.scope, coin_type)
            self._render_all_nodes2()
        # self.end()
        self._excel_factory(path)

    def _personal_note(self, file, scope, coin_type: str):
        self._main_address = ""
        with open(file, newline='') as f:
            self.rf = json.loads(f.read())
            # render_node = True
            if "graph_dic" not in self.rf:
                return
            text = f"下图显示了所有被 {scope} USDT 或以下忽略的记录交易"
            with self.dot.subgraph(name='legend') as c:
                c.attr(color='blue')
                c.node_attr['style'] = 'filled'
                c.attr(label=text)

            nodes = self.rf["graph_dic"]["node_list"]
            edges = self.rf["graph_dic"]["edge_list"]

            for v in nodes:
                _font_colr = "slateblue4"
                _id = v["id"]
                _add = v["addr"]
                _lab = ""

                if "shape" in v:
                    _shape = v["shape"]
                    if "star" in _shape:
                        self._main_address = _id
                else:
                    _shape = "box"

                if _shape == "circularImage":
                    _shape = "box"

                if _add not in self.metadata:
                    self.metadata[_add] = _id
                else:
                    old_id = self.metadata[_add]
                    self.metadata["LINK"][_id] = old_id
                    _id = old_id

                    # render_node = False

                if "..." in v["label"]:
                    _lab = v["addr"]
                    _color = "springgreen2"
                    side_note = self.api.overview(_add, coin_type)
                    _lab = _lab + side_note
                    _font_colr = "white"
                    _shape = "folder"

                if "color" in v:
                    _color = v["color"]

                self.metadata["IDS"][_id] = {
                    "shape": _shape,
                    "fillcolor": _color,
                    "style": "filled",
                    "fontcolor": _font_colr,
                    "label": _lab,
                    "address": v["addr"],
                    "thickness": '1'
                }

            self._weighted_thickness(edges)
            self._final_sort(edges, scope, coin_type, self.extra_func_personalize_note)

    def _weighted_thickness(self, raw_nodes):
        if self._wire_thickness_requirement is True and len(raw_nodes) > 10:
            listnumber = [float(s["val"]) for s in raw_nodes]
            mean = statistics.mean(listnumber)
            # Calculate the squared differences
            squared_differences = [(x - mean) ** 2 for x in listnumber]
            # Calculate the sum of the squared differences
            sum_of_squared_differences = sum(squared_differences)
            # Calculate the number of numbers in the list
            n = len(listnumber)
            # Calculate the standard deviation
            ef = sum_of_squared_differences / (n - 1)
            standard_deviation = mean / 3
            self.metadata["MEAN"] = mean
            self.metadata["STD"] = standard_deviation
            print("mean", mean)
            print("std", standard_deviation)
            s2 = standard_deviation * 2
            s3 = standard_deviation * 3
            s1 = standard_deviation
            for g in raw_nodes:
                _id = self.getId(g["from"])

                _val = g["val"]
                if _val > mean:
                    if _val > mean + s3:
                        self.metadata["IDS"][_id]["thickness"] = '8'
                    elif _val > mean + s2:
                        self.metadata["IDS"][_id]["thickness"] = '7'
                    elif _val > mean + s1:
                        self.metadata["IDS"][_id]["thickness"] = '6'
                    else:
                        self.metadata["IDS"][_id]["thickness"] = '5'
                else:
                    if _val < mean - s3:
                        self.metadata["IDS"][_id]["thickness"] = '1'
                    elif _val < mean - s2:
                        self.metadata["IDS"][_id]["thickness"] = '2'
                    elif _val < mean - s1:
                        self.metadata["IDS"][_id]["thickness"] = '3'
                    else:
                        self.metadata["IDS"][_id]["thickness"] = '4'

    def _inside(self, dict_map: dict, scope: int, coin_type: str = None):
        if "graph_dic" not in dict_map:
            print(dict_map)
            print("-------------- error from this file... ")
            return

        text = f"下图显示了所有被 {scope} USDT 或以下忽略的记录交易"

        with self.dot.subgraph(name='legend') as c:
            c.attr(color='blue')
            c.node_attr['style'] = 'filled'
            c.attr(label=text)

        nodes = dict_map["graph_dic"]["node_list"]
        edges = dict_map["graph_dic"]["edge_list"]
        print("start drawing the plots")
        for v in nodes:
            _font_colr = "black"
            _id = v["id"]
            _add = v["addr"]
            special_address = False

            if "shape" in v:
                _shape = v["shape"]
            else:
                _shape = "box"

            if _shape == "circularImage":
                _shape = "box"

            if _add not in self.metadata:
                self.metadata[_add] = _id
            else:
                old_id = self.metadata[_add]
                self.metadata["LINK"][_id] = old_id
                _id = old_id
                # render_node = False

            if "..." in v["label"]:
                _lab = v["addr"]
                _color = "lightgreen"
                if "star" in _shape:
                    side_note = self.api.overview(_add, coin_type)
                    _lab = _lab + side_note
                    _font_colr = "white"
                    _shape = "folder"
                    _color = "darkslategrey"
                    special_address = True
            else:
                _color = "lightcoral"
                _lab = f"[{v['label']}]\n{v['addr']}"
                if "star" in _shape:
                    side_note = self.api.overview(_add, coin_type)
                    _lab = _lab + side_note
                    _font_colr = "white"
                    _shape = "folder"
                    _color = "darkslategrey"
                    special_address = True

                if self.enable_sidenote:
                    side_note = self.api.overview(_add, coin_type)
                    _lab = _lab + side_note

            # if "color" in v:
            #    _color = v["color"]

            self.metadata["IDS"][_id] = {
                "shape": _shape,
                "fillcolor": _color,
                "style": "filled",
                "fontcolor": _font_colr,
                "label": _lab,
                "thickness": "1",
                "address": _add,
                "common_address": not special_address
            }
        self._weighted_thickness(edges)
        self._final_sort(edges, self.scope, coin_type)

    def _processing_layer(self, layer_addresses: list):
        # processing the layer
        for a in layer_addresses:
            file_path = os.path.join(self.mist_folder, f"{a}.json")
            if os.path.exists(file_path) is False:
                self.api.save(a)
            else:
                self.api.save_map(file_path, a)

    def _render_all_nodes(self):
        for _id in self.metadata["USE_NODE"]:
            # if render_node:
            ob = self.metadata["IDS"][_id]
            self.dot.node(
                _id,
                shape=ob["shape"],
                fillcolor=ob["fillcolor"],
                style="filled",
                fontcolor=ob["fontcolor"],
                label=ob["label"]
            )

    def _render_all_nodes2(self):
        for _id in self.metadata["USE_NODE"]:
            # if render_node:
            ob = self.metadata["IDS"][_id]
            self.dot.node(
                _id,
                shape=ob["shape"],
                fillcolor=ob["fillcolor"],
                style="filled",
                fontcolor=ob["fontcolor"],
                label=ob["label"],
                constraint="false"
            )

    def mark_next_layer(self, from_address_id, to_address_id):
        f_id = self.getId(from_address_id)
        t_id = self.getId(to_address_id)
        from_address = self.metadata["IDS"][f_id]["address"]
        c1 = self.metadata["IDS"][f_id]["common_address"]
        to_address = self.metadata["IDS"][t_id]["address"]
        c2 = self.metadata["IDS"][t_id]["common_address"]
        self.layer.dump_layer_addresses(from_address, to_address, c1, c2)

    def _final_sort(self, edges_nodes_dict: list, scope: int, coin_type: str, extra_func=None):
        """
        render the final graphivz image with the given configs
        """
        for v in edges_nodes_dict:
            if v["val"] < scope:
                continue

            if "color" in v:
                _color = v["color"]["color"]
            else:
                _color = "red"
            count = len(v["tx_hash_list"])
            __label = v["label"]
            self.dot.edge(
                self.getId(v["from"]),
                self.getId(v["to"]),
                color=_color,
                label=f"{count}筆,{__label}",
                labeldistance='1.2',
                labelangle='60',
                penwidth=self.getThickness(v["from"])
            )
            self.mark_next_layer(v["from"], v["to"])
            if extra_func is not None:
                extra_func(v, count, __label, coin_type)
            self.edges += 1

    def extra_func_personalize_note(self, item_note: dict, spend_count: int, label_contribution: str, coin_type: str):
        """
        extra personalized note for personal note only.
        """
        if self._main_address == self.getId(item_note["to"]) and self.use_from:
            from_id = self.getId(item_note["from"])
            address_from = self.metadata["IDS"][from_id]["address"]
            # print("record -> excel")
            self.metadata["FROM_LIST"].append({
                "info": self.api.overview_n_dict(address_from, coin_type),
                "address": address_from,
                "spent_count": spend_count,
                "contribute": label_contribution
            })

        if self._main_address == self.getId(item_note["from"]) and self.use_to:
            from_id = self.getId(item_note["to"])
            address_from = self.metadata["IDS"][from_id]["address"]
            # print("record -> excel")
            self.metadata["FROM_LIST"].append({
                "info": self.api.overview_n_dict(address_from, coin_type),
                "address": address_from,
                "spent_count": spend_count,
                "contribute": label_contribution
            })

    def getId(self, _idx):
        if _idx in self.metadata["LINK"]:
            use_alternative = self.metadata["LINK"][_idx]
        else:
            use_alternative = _idx

        if use_alternative not in self.metadata["USE_NODE"]:
            self.metadata["USE_NODE"].append(use_alternative)

        return use_alternative

    def getThickness(self, _id_index) -> int:
        if self._wire_thickness_requirement is False:
            return 1
        else:
            code_id = self.getId(_id_index)
            return self.metadata["IDS"][code_id]["thickness"]

    def end(self):
        self.dot.render(
            filename=f"{self.project_name}.{self.scope}",
            directory=self.chart_folder).replace('\\', '/')

    def end_with_k(self, k: int):
        self.dot.render(
            filename=f"{self.project_name}.{self.scope}.{k}",
            directory=self.chart_folder).replace('\\', '/')

    def setName(self, name: str):
        self.project_name = name

    def _readDF(self, excel_file: str) -> DataFrame:
        data = pd.read_excel(excel_file)
        df = pd.DataFrame(data, columns=self._excel_header)
        return df

    def _excel_factory(self, excel_file: str) -> "MistAnalysis":

        self._excel_header = [
            # 2         3         4       5        6         7     8          9            10
            "address", "start", "end", "income", "outflow", "net", "tx", "spent count", "contribution"
        ]
        sheet_name = ""

        if self.use_to is True:
            sheet_name = "income"

        if self.use_from is True:
            sheet_name = "expense"

        try:
            _df = self._readDF(excel_file)
            wb = ox.load_workbook(excel_file)
            ws = wb[sheet_name]
            startrow: int = 0
            original_rows = len(self.metadata["FROM_LIST"])
            for row in range(0, original_rows):
                if row >= startrow:
                    rT = row + 2
                    package_info = self.metadata["FROM_LIST"][row]["info"]
                    address = self.metadata["FROM_LIST"][row]["address"]
                    spent_count = self.metadata["FROM_LIST"][row]["spent_count"]
                    contribute = self.metadata["FROM_LIST"][row]["contribute"]
                    # for ch
                    ws.cell(row=rT, column=1).value = address
                    ws.cell(row=rT, column=2).value = package_info["first_tx_time"]
                    ws.cell(row=rT, column=3).value = package_info["last_tx_time"]
                    ws.cell(row=rT, column=4).value = package_info["total_received"]
                    ws.cell(row=rT, column=5).value = package_info["total_spent"]
                    ws.cell(row=rT, column=6).value = package_info["balance"]
                    ws.cell(row=rT, column=7).value = package_info["tx_count"]
                    ws.cell(row=rT, column=8).value = spent_count
                    ws.cell(row=rT, column=9).value = contribute
                    print(address)
                    # for ch in range(1, original_cols + 1):
                    # ws.cell(row=rT, column=ch).fill = pf
                    #    pass

            """
            for row in range(0, original_rows):  # For each row in the dataframe
                for col in range(0, original_cols):  # For each column in the dataframe
                    if row >= startrow:
                        val_old = _df.iat[row, col]
                        key = self._excel_header[col]
                        if col == 0:
                            rT = row + 2
                            ws.cell(row=rT, column=2).value = pass_discord
                            ws.cell(row=rT, column=4).value = pass_email
                            ws.cell(row=rT, column=5).value = token_p
                            ws.cell(row=rT, column=9).value = banned
                            ws.cell(row=rT, column=10).value = last_operation
                            ws.cell(row=rT, column=11).value = bind_phone
                            # print(f"Pop =|> {rT} {val_old} {k1v1} {k1v2} {k1v3}")
            """

            wb.save(excel_file)

        except FileNotFoundError:
            pass

        return self

    def lineThickness(self, auto: bool = True):
        self._wire_thickness_requirement = auto
        return self

    def acquireFromAddress(self, addresses: list):
        for x in addresses:
            self.api.save(x)
        self.layer.add_first_layer(addresses)

    def acquireByFile(self, file_name: str, auto_sep: bool = False):
        path = os.path.join(self.inputfolder, file_name)
        if os.path.isfile(path) is False:
            print("The address file is not found under the inputs folder.")
            exit(2)
        addresses = readALlLines(path)
        for x in addresses:
            if auto_sep is True:
                if x[:2] == "0x":
                    enable_USDT_ERC20()
                    self.api.save(x, True)
                    enable_USDT_ARB()
                    self.api.save(x, True)
                    enable_DAI_POLYGON()
                    self.api.save(x, True)
                    enable_USDT_BEP20()
                    self.api.save(x, True)

                elif x[:1] == "T":
                    enable_USDT_TRC20()
                    self.api.save(x, True)
            else:
                # default use erc20
                if x[:2] == "0x":
                    self.api.save(x, True)
        count_addr = len(addresses)
        print(f"All {count_addr} addresses added to the first layer")
        self.layer.add_first_layer(addresses)


class SlowMistAnalysisV2(MistAnalysis):
    def startPlotV2SingleOneLayer(self):
        """
        will be using the db data to make plots
        """
        set_coin_type = "xxx"
        addresses = self.layer.get_current_layer_address()

        if self.is_independence is False:
            while True:
                for address_x in addresses:
                    list_reports = self.api.sql.get_map_by_address(address_x)
                    if len(list_reports) == 0:
                        print(f"report for {address_x} is not found")
                        continue
                    for xx in list_reports:
                        (address, file, coin_type, updatetime) = xx
                        if set_coin_type == coin_type:
                            data_raw = json.loads(file)
                            self._inside(data_raw, self.scope, coin_type)
                            break

                    if self.layer.next_depth() is False:
                        break
        else:
            y = 0
            for address_x in addresses:
                self.reset_metadata()
                self.dot = build_bot_tpl(
                    "collection",
                    self.scope
                )
                list_reports = self.api.sql.get_map_by_address(address_x)
                if len(list_reports) == 0:
                    print(f"report for {address_x} is not found")
                    continue
                for xx in list_reports:
                    (address, file, coin_type, updatetime) = xx

                    data_raw = json.loads(file)
                    self._inside(data_raw, self.scope, coin_type)
                    self._render_all_nodes()
                    self.end_with_k(y)

                    y += 1
